#!/usr/bin/python
# -*- coding: UTF-8 -*-
'''
Author：BetterDefender
Version：1.0.1

For Example
python Excel_Filter.py inputfile keyword outfile
sys.argv[0]=Excel_Filter.py    0为python后的第一个项也就是 Excel_Filter.py
sys.argv[1]=inputfile  输入文件
sys.argv[2]=login   关键字
sys.argv[3]=outfile 此项为输出文件
'''
import sys
import numpy as np
import pandas as pd
import os
import xlrd
import xlwt
from xlutils3.copy import copy
import openpyxl

def readtxt(infile,key,outfile):    #txt文本提取方法
    with open(outfile, 'r') as file_to_read:
        s = file_to_read.readlines()

    with open(infile, 'r') as file_to_read:
        while True:
            line = file_to_read.readline()  # 整行读取数据
            if not line:
                break
                pass
            if key in line:
                    if s:
                        out = open(outfile, 'a+')
                        out.write(str(line))  # 保存入结果文件 写入w / 追加a+
                        #out.write('\n')  # 保存入结果文件 写入w / 追加a+
                    else:
                        open(outfile, 'w').write(str(line))  # 保存入结果文件 写入w / 追加a+
                        s = 1
    print('提取完成，请查看输出文件',outfile)

def readxlsx(infile,key,outfile):   #xls或xlsx提取方法

    data = xlrd.open_workbook(infile)  # xlrd读取xls\xlsx都可以
    # print(data.sheet_names()) # 输出所有页的名称
    table = data.sheets()[0]  # 获取第一页
    # table = data.sheet_by_index(0) # 通过索引获得第一页
    # table = data.sheet_by_name('Sheet1') # 通过名称来获取指定页
    nrows = table.nrows  # 为行数，整形
    ncolumns = table.ncols  # 为列数，整形
    # print(type(nrows))
    # print(table.row_values(0))# 输出第一行值 为一个列表
    # 遍历输出所有行值

    if outfile.endswith('.xls'):
        excel_data = xlrd.open_workbook(outfile)
        excel = copy(wb=excel_data)
        excel_table = excel.get_sheet(0)  # 获得要操作的页
        nrows2 = excel_data.sheets()[0].nrows  # 获得行数
        for row in range(nrows):
            for value in table.row_values(row):
                if key in str(value):
                    ncols2 = 0
                    for value in table.row_values(row):
                        excel_table.write(nrows2, ncols2, value)
                        ncols2 += 1
                    nrows2 += 1
            continue
        excel.save(outfile)
        print('提取完成，请查看输出文件',outfile)

    if outfile.endswith('.xlsx'):
        out_data = openpyxl.load_workbook(outfile)
        # print(data.get_named_ranges())  # 输出工作页索引范围
        # print(data.get_sheet_names())  # 输出所有工作页的名称
        # 取第一张表
        #out_table = out_data.get_sheet(0)
        out_table = out_data.active
        #print(type(out_table))
        # print(table.title)  # 输出表名
        nrows2 = out_table.max_row  # 获得行数
        #ncols2 = out_table.max_column  # 获得列数

        for row in range(nrows):
            for value in table.row_values(row):
                if key in str(value):
                    ncols2 = 1
                    for value in table.row_values(row):
                        out_table.cell(nrows2, ncols2).value = value
                        ncols2 += 1
                    nrows2 = nrows2 + 1
            continue
        out_data.save(outfile)
        print('提取完成，请查看输出文件',outfile)

print('\n')
print('       _____ _           _')
print('      |  ___(_)_ __   __| |')
print('      | |_  | | |_ \ / _` |')
print('      |  _| | | | | | (_| |')
print('      |_|   |_|_| |_|\__,_|')
print('\n')
print('                      Author：BetterDefender')
print('                      Version：1.0.1')

try:
    #python Excel_Filtere.py inputfile keyword outfile
    file = sys.argv[1] #获取用户输入路径
    keyword = sys.argv[2]   #获取关键字
    outfile = sys.argv[3]   #获取输出文件

    if __name__ == '__main__': 
        #ile = 'abc.txt'
        #outfile = 'temp.txt'

        #file = 'target.xlsx'
        #outfile = 'excel_test.xls'

        #file = 'target.xlsx'
        #outfile = 'excel_table.xlsx'

        #file = 'target.xls'
        #outfile = 'excel_table.xlsx'

        #file = 'target.xls'
        #outfile = 'excel_test.xls'

        #判断用户传入文本类型，暂时只支持txt和xlsx和xls格式的提取
        if file.endswith('.xlsx') or file.endswith('.xls'):
            readxlsx(file, keyword, outfile)
        elif file.endswith('.txt'):
            readtxt(file, keyword, outfile)
        else:
            print('------------------------')
            print('\n')
            print('Tips：只支持txt和xls和xlsx类型')
            print('\n')
            print('------------------------')
except:
    print('\n')
    print('------------------------------------------')
    print('\n')
    print('For Example:')
    print('python Excel_Filtere.py inputfile keyword outfile')
    print('\n')
    print('------------------------------------------')